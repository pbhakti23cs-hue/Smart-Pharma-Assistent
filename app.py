from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, g, make_response, send_file
import sqlite3
import joblib
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import os
import threading
import time
import traceback
import csv
import io
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import tempfile
from PIL import Image
import matplotlib.pyplot as plt
import numpy as np

app = Flask(__name__)
app.secret_key = 'smart-pharma-assistant-secret-key-2024'

# Load ML model
try:
    model_path = 'models/symptom_model.joblib'
    if os.path.exists(model_path):
        symptom_model = joblib.load(model_path)
        print("✅ ML Model loaded successfully")
    else:
        print("⚠️  ML Model file not found. Please run train_model.py first")
        symptom_model = None
except Exception as e:
    print(f"⚠️  Error loading ML model: {str(e)}")
    symptom_model = None

# Database helper
def get_db():
    conn = sqlite3.connect('pharma.db')
    conn.row_factory = sqlite3.Row
    return conn

# Initialize database
def init_database():
    try:
        db = get_db()
        db.execute('SELECT 1 FROM users LIMIT 1')
        db.close()
        print("✅ Database already initialized")
    except sqlite3.OperationalError:
        print("Database tables not found. Running initialization...")
        try:
            exec(open('init_db.py').read())
        except Exception as e:
            print(f"Error initializing database: {e}")

# Initialize database on startup
init_database()

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login first', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Function to create alerts
def create_alert(alert_type, message, medicine_id=None, batch_id=None, priority='medium'):
    try:
        db = get_db()
        # Check if similar alert already exists
        existing = db.execute('''
            SELECT id FROM alerts 
            WHERE alert_type = ? AND message LIKE ? AND is_read = 0
            LIMIT 1
        ''', (alert_type, f'%{message[:50]}%')).fetchone()
        
        if not existing:
            db.execute('''
                INSERT INTO alerts (alert_type, message, medicine_id, batch_id, priority, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (alert_type, message, medicine_id, batch_id, priority, datetime.now()))
            db.commit()
        db.close()
        return True
    except Exception as e:
        print(f"Error creating alert: {e}")
        return False

# Function to check and create alerts automatically
def check_and_create_alerts():
    try:
        db = get_db()
        
        # 1. Check for low stock medicines (less than 20 units total)
        medicines = db.execute('''
            SELECT m.id, m.name, 
                   COALESCE(SUM(b.quantity), 0) as total_stock
            FROM medicines m
            LEFT JOIN batches b ON m.id = b.medicine_id
            GROUP BY m.id, m.name
            HAVING total_stock > 0
        ''').fetchall()
        
        for medicine in medicines:
            total_stock = medicine['total_stock']
            if total_stock <= 20:
                create_alert(
                    alert_type='low_stock',
                    message=f'{medicine["name"]} is running low ({total_stock} units left)',
                    medicine_id=medicine['id'],
                    priority='high' if total_stock <= 5 else 'medium'
                )
        
        # 2. Check for near expiry batches (within 30 days)
        near_expiry_batches = db.execute('''
            SELECT b.id, b.batch_no, b.expiry_date, m.name as medicine_name, m.id as medicine_id
            FROM batches b
            JOIN medicines m ON b.medicine_id = m.id
            WHERE b.expiry_date BETWEEN DATE('now') AND DATE('now', '+30 days')
            AND b.quantity > 0
        ''').fetchall()
        
        for batch in near_expiry_batches:
            expiry_date = datetime.strptime(batch['expiry_date'], '%Y-%m-%d').date()
            days_left = (expiry_date - datetime.now().date()).days
            create_alert(
                alert_type='expiry',
                message=f'Batch {batch["batch_no"]} of {batch["medicine_name"]} expires in {days_left} days',
                medicine_id=batch['medicine_id'],
                batch_id=batch['id'],
                priority='high' if days_left <= 7 else 'medium'
            )
        
        # 3. Check for expired batches
        expired_batches = db.execute('''
            SELECT b.id, b.batch_no, m.name as medicine_name, m.id as medicine_id
            FROM batches b
            JOIN medicines m ON b.medicine_id = m.id
            WHERE b.expiry_date < DATE('now') AND b.quantity > 0
        ''').fetchall()
        
        for batch in expired_batches:
            create_alert(
                alert_type='expired',
                message=f'Batch {batch["batch_no"]} of {batch["medicine_name"]} has expired',
                medicine_id=batch['medicine_id'],
                batch_id=batch['id'],
                priority='high'
            )
        
        db.close()
        return True
    except Exception as e:
        print(f"Error checking alerts: {e}")
        return False

# Background alert checker thread
def alert_checker_thread():
    """Background thread to check alerts periodically"""
    while True:
        try:
            check_and_create_alerts()
            time.sleep(300)  # Check every 5 minutes
        except Exception as e:
            print(f"Error in alert checker thread: {e}")
            time.sleep(60)

# Start alert checker thread
def start_alert_checker():
    if not hasattr(app, 'alert_checker_started'):
        thread = threading.Thread(target=alert_checker_thread, daemon=True)
        thread.start()
        app.alert_checker_started = True
        print("✅ Alert checker thread started")

# Before each request - add unread alerts count to g
@app.before_request
def before_request():
    if 'user_id' in session:
        db = get_db()
        unread_alerts = db.execute('SELECT COUNT(*) as count FROM alerts WHERE is_read = 0').fetchone()['count']
        g.unread_alerts_count = unread_alerts
        db.close()

# Routes
@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            flash('❌ Please fill all fields', 'danger')
            return redirect(url_for('login'))
        
        try:
            db = get_db()
            user = db.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
            db.close()
            
            if user and check_password_hash(user['password_hash'], password):
                session['user_id'] = user['id']
                session['username'] = user['username']
                
                # Start alert checker on first login
                start_alert_checker()
                
                flash('✅ Login successful!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('❌ Invalid username or password', 'danger')
        except Exception as e:
            flash(f'❌ Database error: {str(e)}', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('👋 Logged out successfully', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    try:
        db = get_db()
        
        # Get statistics
        stats = {
            'medicines': db.execute('SELECT COUNT(*) FROM medicines').fetchone()[0] or 0,
            'batches': db.execute('SELECT COUNT(*) FROM batches').fetchone()[0] or 0,
            'sales': db.execute('SELECT SUM(quantity_sold) FROM sales').fetchone()[0] or 0,
            'revenue': db.execute('SELECT SUM(quantity_sold * selling_price) FROM sales').fetchone()[0] or 0,
        }
        
        # Near expiry batches (within 15 days)
        near_expiry = db.execute('''
            SELECT COUNT(*) FROM batches 
            WHERE expiry_date BETWEEN DATE('now') AND DATE('now', '+15 days')
            AND quantity > 0
        ''').fetchone()[0] or 0
        
        # Expired batches
        expired = db.execute('SELECT COUNT(*) FROM batches WHERE expiry_date < DATE("now") AND quantity > 0').fetchone()[0] or 0
        
        # Low stock medicines (< 20 units)
        low_stock = db.execute('''
            SELECT COUNT(*) FROM (
                SELECT m.id, COALESCE(SUM(b.quantity), 0) as total
                FROM medicines m
                LEFT JOIN batches b ON m.id = b.medicine_id
                GROUP BY m.id
                HAVING total <= 20 AND total > 0
            )
        ''').fetchone()[0] or 0
        
        # Get recent alerts
        recent_alerts = db.execute('''
            SELECT a.*, m.name as medicine_name, b.batch_no
            FROM alerts a
            LEFT JOIN medicines m ON a.medicine_id = m.id
            LEFT JOIN batches b ON a.batch_id = b.id
            WHERE a.is_read = 0
            ORDER BY a.created_at DESC
            LIMIT 5
        ''').fetchall()
        
        db.close()
        
        now = datetime.now()
        return render_template('dashboard.html',
                             stats=stats,
                             near_expiry=near_expiry,
                             expired=expired,
                             low_stock=low_stock,
                             recent_alerts=recent_alerts,
                             username=session['username'],
                             current_date=now,
                             current_time=now)
    except Exception as e:
        flash(f'❌ Error loading dashboard: {str(e)}', 'danger')
        return redirect(url_for('login'))

@app.route('/medicines')
@login_required
def view_medicines():
    try:
        db = get_db()
        medicines = db.execute('''
            SELECT m.*, 
                   COALESCE(SUM(b.quantity), 0) as total_stock,
                   COUNT(b.id) as batch_count
            FROM medicines m
            LEFT JOIN batches b ON m.id = b.medicine_id
            GROUP BY m.id
            ORDER BY m.name
        ''').fetchall()
        
        db.close()
        return render_template('view_medicines.html', 
                             medicines=medicines,
                             current_date=datetime.now())
    except Exception as e:
        flash(f'❌ Error loading medicines: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/medicine/add', methods=['GET', 'POST'])
@login_required
def add_medicine():
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            composition = request.form.get('composition', '').strip()
            uses = request.form.get('uses', '').strip()
            dosage = request.form.get('dosage', '').strip()
            side_effects = request.form.get('side_effects', '').strip()
            category = request.form.get('category', '').strip()
            
            if not name:
                flash('❌ Medicine name is required', 'danger')
                return redirect(url_for('add_medicine'))
            
            db = get_db()
            db.execute('''
                INSERT INTO medicines (name, composition, uses, dosage, side_effects, category, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (name, composition, uses, dosage, side_effects, category, datetime.now()))
            db.commit()
            db.close()
            
            flash(f'✅ Medicine "{name}" added successfully!', 'success')
            return redirect(url_for('view_medicines'))
        except Exception as e:
            flash(f'❌ Error adding medicine: {str(e)}', 'danger')
            return redirect(url_for('add_medicine'))
    
    return render_template('add_medicine.html', current_date=datetime.now())

@app.route('/medicine/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_medicine(id):
    try:
        db = get_db()
        
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            composition = request.form.get('composition', '').strip()
            uses = request.form.get('uses', '').strip()
            dosage = request.form.get('dosage', '').strip()
            side_effects = request.form.get('side_effects', '').strip()
            category = request.form.get('category', '').strip()
            
            if not name:
                flash('❌ Medicine name is required', 'danger')
                return redirect(url_for('edit_medicine', id=id))
            
            db.execute('''
                UPDATE medicines 
                SET name = ?, composition = ?, uses = ?, dosage = ?, 
                    side_effects = ?, category = ?, updated_at = ?
                WHERE id = ?
            ''', (name, composition, uses, dosage, side_effects, category, datetime.now(), id))
            db.commit()
            
            flash(f'✅ Medicine "{name}" updated successfully!', 'success')
            return redirect(url_for('view_medicines'))
        
        medicine = db.execute('SELECT * FROM medicines WHERE id = ?', (id,)).fetchone()
        db.close()
        
        if not medicine:
            flash('❌ Medicine not found', 'danger')
            return redirect(url_for('view_medicines'))
        
        return render_template('edit_medicine.html',
                             medicine=medicine,
                             current_date=datetime.now())
    
    except Exception as e:
        flash(f'❌ Error editing medicine: {str(e)}', 'danger')
        return redirect(url_for('view_medicines'))

@app.route('/medicine/delete/<int:id>')
@login_required
def delete_medicine(id):
    try:
        db = get_db()
        medicine = db.execute('SELECT name FROM medicines WHERE id = ?', (id,)).fetchone()
        
        db.execute('DELETE FROM medicines WHERE id = ?', (id,))
        db.commit()
        db.close()
        
        if medicine:
            flash(f'✅ Medicine "{medicine["name"]}" deleted successfully!', 'success')
        else:
            flash('✅ Medicine deleted successfully!', 'success')
        
        return redirect(url_for('view_medicines'))
    
    except Exception as e:
        flash(f'❌ Error deleting medicine: {str(e)}', 'danger')
        return redirect(url_for('view_medicines'))

@app.route('/batch/add', methods=['GET', 'POST'])
@login_required
def add_batch():
    try:
        db = get_db()
        
        if request.method == 'POST':
            try:
                medicine_id = request.form.get('medicine_id')
                batch_no = request.form.get('batch_no', '').strip()
                quantity = int(request.form.get('quantity', '0'))
                mrp = float(request.form.get('mrp', '0.0'))
                cost_price = float(request.form.get('cost_price', '0.0'))
                mfg_date = request.form.get('mfg_date')
                expiry_date = request.form.get('expiry_date')
                supplier = request.form.get('supplier', '').strip()
                
                if not batch_no or not medicine_id:
                    flash('❌ Batch number and medicine are required', 'danger')
                    return redirect(url_for('add_batch'))
                
                # Check if batch number already exists
                existing = db.execute('SELECT id FROM batches WHERE batch_no = ?', (batch_no,)).fetchone()
                if existing:
                    flash('❌ Batch number already exists', 'danger')
                    return redirect(url_for('add_batch'))
                
                db.execute('''
                    INSERT INTO batches (medicine_id, batch_no, quantity, mrp, cost_price, 
                                        mfg_date, expiry_date, supplier, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (medicine_id, batch_no, quantity, mrp, cost_price, 
                      mfg_date, expiry_date, supplier, datetime.now()))
                db.commit()
                
                # Create alert for new batch with near expiry
                if expiry_date:
                    expiry_date_obj = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                    days_until_expiry = (expiry_date_obj - datetime.now().date()).days
                    if days_until_expiry <= 30:
                        medicine_name = db.execute('SELECT name FROM medicines WHERE id = ?', 
                                                  (medicine_id,)).fetchone()['name']
                        create_alert(
                            alert_type='expiry',
                            message=f'New batch {batch_no} of {medicine_name} expires in {days_until_expiry} days',
                            medicine_id=medicine_id,
                            priority='high' if days_until_expiry <= 7 else 'medium'
                        )
                
                db.close()
                flash(f'✅ Batch "{batch_no}" added successfully!', 'success')
                return redirect(url_for('view_medicines'))
            except ValueError:
                flash('❌ Invalid numeric values', 'danger')
                return redirect(url_for('add_batch'))
            except Exception as e:
                flash(f'❌ Error adding batch: {str(e)}', 'danger')
                return redirect(url_for('add_batch'))
        
        medicines = db.execute('SELECT id, name FROM medicines ORDER BY name').fetchall()
        db.close()
        return render_template('add_batch.html', 
                             medicines=medicines,
                             current_date=datetime.now())
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/sell', methods=['GET', 'POST'])
@login_required
def sell_medicine():
    try:
        db = get_db()
        
        if request.method == 'POST':
            print("\n" + "="*80)
            print("DEBUG: /sell POST request received")
            print("="*80)
            
            try:
                # Get customer info
                customer_name = request.form.get('customer_name', '').strip()
                customer_phone = request.form.get('customer_phone', '').strip()
                customer_age = request.form.get('customer_age', '')
                prescription_number = request.form.get('prescription_number', '').strip()
                doctor_name = request.form.get('doctor_name', '').strip()
                diagnosis = request.form.get('diagnosis', '').strip()
                payment_method = request.form.get('payment_method', 'cash')
                
                # Get cart items
                batch_ids = request.form.getlist('batch_id[]')
                quantities = request.form.getlist('quantity[]')
                prices = request.form.getlist('price[]')
                
                print(f"DEBUG: Customer: {customer_name}")
                print(f"DEBUG: Phone: {customer_phone}")
                print(f"DEBUG: Payment: {payment_method}")
                print(f"DEBUG: Batch IDs: {batch_ids}")
                print(f"DEBUG: Quantities: {quantities}")
                print(f"DEBUG: Prices: {prices}")
                
                # Validate required fields
                if not customer_name or not customer_phone:
                    error_msg = 'Customer name and phone are required'
                    print(f"DEBUG: Validation failed - {error_msg}")
                    return jsonify({
                        'success': False,
                        'error': error_msg
                    }), 400
                
                if not batch_ids or len(batch_ids) == 0:
                    error_msg = 'Please add medicines to cart'
                    print(f"DEBUG: Validation failed - {error_msg}")
                    return jsonify({
                        'success': False,
                        'error': error_msg
                    }), 400
                
                # Check arrays have same length
                if not (len(batch_ids) == len(quantities) == len(prices)):
                    error_msg = f"Cart data mismatch"
                    print(f"DEBUG: Validation failed - {error_msg}")
                    return jsonify({
                        'success': False,
                        'error': error_msg
                    }), 400
                
                print(f"DEBUG: Processing {len(batch_ids)} cart items...")
                
                # Process each item
                sale_items = []
                total_amount = 0
                
                for i in range(len(batch_ids)):
                    try:
                        batch_id = int(batch_ids[i])
                        quantity = int(quantities[i])
                        price = float(prices[i])
                        
                        print(f"DEBUG: Item {i+1} - Batch ID: {batch_id}, Qty: {quantity}, Price: {price}")
                        
                        # Check stock
                        batch = db.execute('SELECT * FROM batches WHERE id = ?', (batch_id,)).fetchone()
                        if not batch:
                            error_msg = f'Batch not found: {batch_id}'
                            print(f"DEBUG: Error - {error_msg}")
                            return jsonify({
                                'success': False,
                                'error': error_msg
                            }), 400
                        
                        print(f"DEBUG: Batch found: {batch['batch_no']}, Stock: {batch['quantity']}")
                        
                        if batch['quantity'] < quantity:
                            error_msg = f'Insufficient stock for batch {batch["batch_no"]}. Available: {batch["quantity"]}, Requested: {quantity}'
                            print(f"DEBUG: Error - {error_msg}")
                            return jsonify({
                                'success': False,
                                'error': error_msg
                            }), 400
                        
                        # Get medicine name
                        medicine = db.execute('SELECT name FROM medicines WHERE id = ?', 
                                            (batch['medicine_id'],)).fetchone()
                        
                        # Add to sale items
                        item_total = price * quantity
                        total_amount += item_total
                        sale_items.append({
                            'medicine_name': medicine['name'] if medicine else 'Unknown',
                            'batch_no': batch['batch_no'],
                            'quantity': quantity,
                            'price': price,
                            'total': item_total
                        })
                        
                        # Update stock
                        db.execute('UPDATE batches SET quantity = quantity - ? WHERE id = ?', 
                                  (quantity, batch_id))
                        
                        # Record sale
                        db.execute('''
                            INSERT INTO sales (batch_id, quantity_sold, selling_price, 
                                             customer_name, customer_phone, customer_age,
                                             prescription_number, doctor_name, diagnosis, 
                                             payment_method, sold_on)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (batch_id, quantity, price, customer_name, customer_phone, 
                              customer_age, prescription_number, doctor_name, diagnosis,
                              payment_method, datetime.now()))
                        
                        print(f"DEBUG: Item {i+1} processed successfully")
                        
                    except Exception as e:
                        error_msg = f'Error processing item {i+1}: {str(e)}'
                        print(f"DEBUG: Error - {error_msg}")
                        return jsonify({
                            'success': False,
                            'error': error_msg
                        }), 400
                
                # Get the last inserted sale ID
                last_sale_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]
                
                # Commit transaction
                db.commit()
                print(f"DEBUG: Transaction committed successfully. Sale ID: {last_sale_id}")
                
                # Calculate totals
                tax_rate = 0.05
                tax_amount = total_amount * tax_rate
                grand_total = total_amount + tax_amount
                
                # Create sale summary
                sale_details = {
                    'sale_id': last_sale_id,
                    'receipt_number': f'SA-{last_sale_id:06d}',
                    'customer_name': customer_name,
                    'customer_phone': customer_phone,
                    'payment_method': payment_method,
                    'sale_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'items': sale_items,
                    'subtotal': round(total_amount, 2),
                    'tax_amount': round(tax_amount, 2),
                    'grand_total': round(grand_total, 2)
                }
                
                db.close()
                
                # Return JSON response
                response_data = {
                    'success': True,
                    'sale_id': last_sale_id,
                    'receipt_number': sale_details['receipt_number'],
                    'sale_details': sale_details,
                    'message': 'Sale processed successfully!'
                }
                
                print(f"DEBUG: Sending JSON response: {response_data}")
                print("="*80)
                
                return jsonify(response_data)
                
            except Exception as e:
                # Rollback on error
                try:
                    db.rollback()
                    print("DEBUG: Transaction rolled back")
                except:
                    pass
                
                error_msg = f'Error processing sale: {str(e)}'
                print(f"DEBUG: Critical error - {error_msg}")
                traceback.print_exc()
                
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), 500
        
        # ========== GET REQUEST ==========
        # GET request - show form
        medicines = db.execute('''
            SELECT m.id, m.name, m.category,
                   b.id as batch_id, b.batch_no, b.quantity, b.mrp, b.expiry_date
            FROM medicines m
            JOIN batches b ON m.id = b.medicine_id
            WHERE b.quantity > 0 AND b.expiry_date > DATE('now')
            ORDER BY m.name, b.expiry_date
        ''').fetchall()
        
        # Generate receipt number
        last_sale = db.execute('SELECT MAX(id) as last_id FROM sales').fetchone()
        receipt_number = (last_sale['last_id'] or 0) + 1
        
        db.close()
        
        # Group medicines
        medicine_dict = {}
        for row in medicines:
            if row['id'] not in medicine_dict:
                medicine_dict[row['id']] = {
                    'id': row['id'],
                    'name': row['name'],
                    'category': row['category'],
                    'batches': []
                }
            
            medicine_dict[row['id']]['batches'].append({
                'id': row['batch_id'],
                'batch_no': row['batch_no'],
                'quantity': row['quantity'],
                'mrp': row['mrp'],
                'expiry_date': row['expiry_date']
            })
        
        medicine_list = list(medicine_dict.values())
        current_time = datetime.now()
        
        return render_template('sell_medicine.html',
                             medicines=medicine_list,
                             receipt_number=receipt_number,
                             current_date=current_time,
                             current_time=current_time)
    
    except Exception as e:
        error_msg = f'Error in sales module: {str(e)}'
        print(f"DEBUG: Unhandled error - {error_msg}")
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'error': error_msg
        }), 500

@app.route('/sales')
@login_required
def view_sales():
    try:
        db = get_db()
        sales = db.execute('''
            SELECT s.*, b.batch_no, m.name as medicine_name,
                   s.quantity_sold * s.selling_price as total_amount
            FROM sales s
            JOIN batches b ON s.batch_id = b.id
            JOIN medicines m ON b.medicine_id = m.id
            ORDER BY s.sold_on DESC
            LIMIT 50
        ''').fetchall()
        
        # Today's summary
        summary = db.execute('''
            SELECT 
                COUNT(*) as total_sales,
                SUM(quantity_sold) as total_units,
                SUM(quantity_sold * selling_price) as total_revenue,
                AVG(quantity_sold * selling_price) as avg_transaction
            FROM sales 
            WHERE DATE(sold_on) = DATE('now')
        ''').fetchone()
        
        db.close()
        
        current_date = datetime.now()
        return render_template('sales.html', 
                             sales=sales, 
                             summary=summary,
                             current_date=current_date)
    except Exception as e:
        flash(f'❌ Error loading sales: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/recommend', methods=['GET', 'POST'])
@login_required
def recommend_medicine():
    if request.method == 'POST':
        symptoms = request.form.get('symptoms', '').strip()
        
        if not symptoms:
            flash('❌ Please enter symptoms', 'danger')
            return redirect(url_for('recommend_medicine'))
        
        if not symptom_model:
            flash('❌ AI model not loaded. Please run train_model.py first.', 'danger')
            return redirect(url_for('recommend_medicine'))
        
        try:
            # Get prediction from ML model
            prediction = symptom_model.predict([symptoms])[0]
            probabilities = symptom_model.predict_proba([symptoms])[0]
            
            # Get top 3 predictions
            classes = symptom_model.classes_
            top_3_indices = probabilities.argsort()[-3:][::-1]
            
            # Get medicine details for recommendations
            recommendations = []
            db = get_db()
            
            for i in top_3_indices:
                medicine_name = classes[i]
                confidence = round(probabilities[i] * 100, 2)
                
                try:
                    # Find the medicine in database to get ID and details
                    medicine = db.execute('''
                        SELECT * FROM medicines 
                        WHERE name LIKE ? LIMIT 1
                    ''', (f'%{medicine_name}%',)).fetchone()
                    
                    if medicine:
                        recommendations.append({
                            'name': medicine['name'],
                            'confidence': confidence,
                            'id': medicine['id'],
                            'details': {
                                'category': medicine['category'],
                                'composition': medicine['composition'],
                                'uses': medicine['uses'],
                                'dosage': medicine['dosage'],
                                'side_effects': medicine['side_effects']
                            }
                        })
                    else:
                        # Medicine not in database, but still show recommendation
                        recommendations.append({
                            'name': medicine_name,
                            'confidence': confidence,
                            'id': None,
                            'details': None
                        })
                        
                except Exception as e:
                    print(f"Error fetching medicine {medicine_name}: {e}")
                    # Still add recommendation without database details
                    recommendations.append({
                        'name': medicine_name,
                        'confidence': confidence,
                        'id': None,
                        'details': None
                    })
            
            db.close()
            
            if not recommendations:
                flash('❌ No recommendations found for these symptoms.', 'warning')
                return render_template('recommend.html', 
                                     symptoms=symptoms,
                                     current_date=datetime.now())
            
            flash(f'✅ Found {len(recommendations)} recommendations for your symptoms.', 'success')
            return render_template('recommend.html',
                                 symptoms=symptoms,
                                 recommendations=recommendations,
                                 current_date=datetime.now())
            
        except Exception as e:
            print(f"Error in AI recommendation: {e}")
            flash(f'❌ Error generating recommendations: {str(e)}', 'danger')
            return render_template('recommend.html', 
                                 current_date=datetime.now())
    
    return render_template('recommend.html', 
                         current_date=datetime.now())

@app.route('/check-interaction', methods=['GET', 'POST'])
@login_required
def check_interaction():
    if request.method == 'POST':
        drug1 = request.form.get('drug1', '').strip()
        drug2 = request.form.get('drug2', '').strip()
        
        if not drug1 or not drug2:
            flash('❌ Please enter both drug names', 'danger')
            return redirect(url_for('check_interaction'))
        
        try:
            db = get_db()
            interaction = db.execute('''
                SELECT * FROM interactions 
                WHERE (drug_a = ? AND drug_b = ?) OR (drug_a = ? AND drug_b = ?)
                LIMIT 1
            ''', (drug1, drug2, drug2, drug1)).fetchone()
            db.close()
            
            return render_template('interaction_result.html',
                                 drug1=drug1,
                                 drug2=drug2,
                                 interaction=interaction,
                                 current_date=datetime.now())
        except Exception as e:
            flash(f'❌ Error checking interaction: {str(e)}', 'danger')
    
    return render_template('check_interaction.html', 
                         current_date=datetime.now())

@app.route('/interaction/result')
@login_required
def interaction_result():
    drug1 = request.args.get('drug1', '')
    drug2 = request.args.get('drug2', '')
    
    try:
        db = get_db()
        interaction = db.execute('''
            SELECT * FROM interactions 
            WHERE (drug_a = ? AND drug_b = ?) OR (drug_a = ? AND drug_b = ?)
            LIMIT 1
        ''', (drug1, drug2, drug2, drug1)).fetchone()
        
        db.close()
        return render_template('interaction_result.html',
                             drug1=drug1,
                             drug2=drug2,
                             interaction=interaction,
                             current_date=datetime.now())
    except Exception as e:
        flash(f'❌ Error loading interaction result: {str(e)}', 'danger')
        return redirect(url_for('check_interaction'))

@app.route('/reports')
@login_required
def reports():
    try:
        db = get_db()
        
        # Sales statistics (last 30 days)
        sales_stats = db.execute('''
            SELECT 
                SUM(quantity_sold * selling_price) as total_revenue,
                SUM(quantity_sold) as total_sales,
                COUNT(DISTINCT customer_name) as unique_customers
            FROM sales 
            WHERE sold_on >= DATE('now', '-30 days')
        ''').fetchone()
        
        # Calculate average transaction value
        avg_transaction = 0
        if sales_stats['total_sales'] and sales_stats['total_revenue']:
            avg_transaction = sales_stats['total_revenue'] / sales_stats['total_sales']
        
        sales_data = {
            'total_revenue': sales_stats['total_revenue'] or 0,
            'total_sales': sales_stats['total_sales'] or 0,
            'unique_customers': sales_stats['unique_customers'] or 0,
            'avg_transaction': avg_transaction
        }
        
        # Inventory statistics
        inventory_stats = db.execute('''
            SELECT 
                COUNT(DISTINCT m.id) as total_medicines,
                COUNT(b.id) as total_batches,
                SUM(b.quantity) as total_quantity,
                SUM(b.quantity * b.mrp) as stock_value,
                SUM(b.quantity * b.cost_price) as cost_value
            FROM medicines m
            LEFT JOIN batches b ON m.id = b.medicine_id
            WHERE b.expiry_date >= DATE('now') OR b.id IS NULL
        ''').fetchone()
        
        # Calculate potential revenue and profit margin
        potential_revenue = inventory_stats['stock_value'] or 0
        cost_value = inventory_stats['cost_value'] or 0
        profit_margin = 0
        if cost_value > 0:
            profit_margin = ((potential_revenue - cost_value) / cost_value) * 100
        
        inventory_data = {
            'total_medicines': inventory_stats['total_medicines'] or 0,
            'total_batches': inventory_stats['total_batches'] or 0,
            'total_quantity': inventory_stats['total_quantity'] or 0,
            'stock_value': inventory_stats['stock_value'] or 0,
            'cost_value': cost_value,
            'potential_revenue': potential_revenue,
            'profit_margin': profit_margin
        }
        
        # Expiry statistics
        expiry_stats = db.execute('''
            SELECT 
                COUNT(CASE WHEN expiry_date < DATE('now') THEN 1 END) as expired_count,
                COUNT(CASE WHEN expiry_date BETWEEN DATE('now') AND DATE('now', '+15 days') THEN 1 END) as near_expiry_count,
                COUNT(CASE WHEN expiry_date BETWEEN DATE('now', '+16 days') AND DATE('now', '+90 days') THEN 1 END) as expiring_soon_count,
                COUNT(CASE WHEN expiry_date > DATE('now', '+90 days') THEN 1 END) as good_stock_count
            FROM batches
        ''').fetchone()
        
        expiry_data = {
            'expired_count': expiry_stats['expired_count'] or 0,
            'near_expiry_count': expiry_stats['near_expiry_count'] or 0,
            'expiring_soon_count': expiry_stats['expiring_soon_count'] or 0,
            'good_stock_count': expiry_stats['good_stock_count'] or 0
        }
        
        db.close()
        
        current_date = datetime.now()
        return render_template('reports.html',
                             sales_stats=sales_data,
                             inventory_stats=inventory_data,
                             expiry_stats=expiry_data,
                             current_date=current_date,
                             now=current_date)
    
    except Exception as e:
        flash(f'❌ Error loading reports: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/alerts')
@login_required
def view_alerts():
    try:
        db = get_db()
        # Check for new alerts first
        check_and_create_alerts()
        
        alerts = db.execute('''
            SELECT a.*, m.name as medicine_name, b.batch_no
            FROM alerts a
            LEFT JOIN medicines m ON a.medicine_id = m.id
            LEFT JOIN batches b ON a.batch_id = b.id
            ORDER BY 
                CASE WHEN a.is_read = 0 THEN 0 ELSE 1 END,
                a.created_at DESC
        ''').fetchall()
        
        db.close()
        return render_template('alerts.html', 
                             alerts=alerts,
                             current_date=datetime.now())
    except Exception as e:
        flash(f'❌ Error loading alerts: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/alerts/mark-read/<int:id>')
@login_required
def mark_alert_read(id):
    try:
        db = get_db()
        db.execute('UPDATE alerts SET is_read = 1 WHERE id = ?', (id,))
        db.commit()
        db.close()
        flash('✅ Alert marked as read', 'success')
    except Exception as e:
        flash(f'❌ Error marking alert as read: {str(e)}', 'danger')
    
    return redirect(url_for('view_alerts'))

@app.route('/alerts/clear-all')
@login_required
def clear_all_alerts():
    try:
        db = get_db()
        db.execute('UPDATE alerts SET is_read = 1 WHERE is_read = 0')
        db.commit()
        db.close()
        flash('✅ All alerts cleared', 'success')
    except Exception as e:
        flash(f'❌ Error clearing alerts: {str(e)}', 'danger')
    
    return redirect(url_for('view_alerts'))

# API Routes for real-time alerts
@app.route('/api/alerts')
@login_required
def get_alerts_api():
    try:
        db = get_db()
        alerts = db.execute('''
            SELECT a.*, m.name as medicine_name, b.batch_no
            FROM alerts a
            LEFT JOIN medicines m ON a.medicine_id = m.id
            LEFT JOIN batches b ON a.batch_id = b.id
            WHERE a.is_read = 0
            ORDER BY a.created_at DESC
            LIMIT 10
        ''').fetchall()
        
        alerts_list = []
        for alert in alerts:
            # Convert datetime to string for JSON
            created_at_str = alert['created_at']
            if isinstance(created_at_str, datetime):
                created_at_str = created_at_str.strftime('%Y-%m-%d %H:%M:%S')
            elif created_at_str:
                created_at_str = str(created_at_str)
            
            alerts_list.append({
                'id': alert['id'],
                'type': alert['alert_type'],
                'message': alert['message'],
                'priority': alert['priority'],
                'created_at': created_at_str,
                'medicine': alert['medicine_name'],
                'batch': alert['batch_no']
            })
        
        db.close()
        return jsonify({
            'success': True,
            'alerts': alerts_list,
            'count': len(alerts_list)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/alerts/mark-read/<int:id>', methods=['POST'])
@login_required
def mark_alert_read_api(id):
    try:
        db = get_db()
        db.execute('UPDATE alerts SET is_read = 1 WHERE id = ?', (id,))
        db.commit()
        db.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Add missing API routes
@app.route('/api/medicines/search')
@login_required
def search_medicines():
    query = request.args.get('q', '').strip()
    
    try:
        db = get_db()
        
        if query:
            medicines = db.execute('''
                SELECT m.*, 
                       (SELECT SUM(quantity) FROM batches b 
                        WHERE b.medicine_id = m.id AND b.expiry_date > DATE('now')) as total_stock
                FROM medicines m
                WHERE m.name LIKE ? OR m.category LIKE ? OR m.composition LIKE ?
                ORDER BY m.name
                LIMIT 20
            ''', (f'%{query}%', f'%{query}%', f'%{query}%')).fetchall()
        else:
            medicines = db.execute('''
                SELECT m.*, 
                       (SELECT SUM(quantity) FROM batches b 
                        WHERE b.medicine_id = m.id AND b.expiry_date > DATE('now')) as total_stock
                FROM medicines m
                ORDER BY m.name
                LIMIT 20
            ''').fetchall()
        
        db.close()
        
        medicines_list = []
        for med in medicines:
            medicines_list.append({
                'id': med['id'],
                'name': med['name'],
                'category': med['category'],
                'composition': med['composition'],
                'total_stock': med['total_stock'] or 0
            })
        
        return jsonify({'medicines': medicines_list})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batches/<int:medicine_id>')
@login_required
def get_batches(medicine_id):
    try:
        print(f"\nDEBUG: Getting batches for medicine_id={medicine_id}")
        
        db = get_db()
        batches = db.execute('''
            SELECT id, batch_no, quantity, mrp, expiry_date
            FROM batches
            WHERE medicine_id = ? AND quantity > 0 AND expiry_date > DATE('now')
            ORDER BY expiry_date
        ''', (medicine_id,)).fetchall()
        
        print(f"DEBUG: Found {len(batches)} batches for medicine {medicine_id}")
        
        db.close()
        
        batches_list = []
        for batch in batches:
            print(f"DEBUG: Batch {batch['id']}: {batch['batch_no']}, Qty: {batch['quantity']}, MRP: {batch['mrp']}")
            batches_list.append({
                'id': batch['id'],
                'batch_no': batch['batch_no'],
                'quantity': batch['quantity'],
                'mrp': batch['mrp'],
                'expiry_date': batch['expiry_date']
            })
        
        return jsonify({'batches': batches_list})
    
    except Exception as e:
        print(f"ERROR in get_batches API: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Add missing routes for medicine details
@app.route('/medicine/details/<int:id>')
@login_required
def medicine_details(id):
    try:
        db = get_db()
        medicine = db.execute('SELECT * FROM medicines WHERE id = ?', (id,)).fetchone()
        
        if not medicine:
            flash('❌ Medicine not found', 'danger')
            return redirect(url_for('view_medicines'))
        
        batches = db.execute('''
            SELECT * FROM batches 
            WHERE medicine_id = ?
            ORDER BY expiry_date
        ''', (id,)).fetchall()
        
        db.close()
        
        return render_template('medicine_details.html',
                             medicine=medicine,
                             batches=batches,
                             current_date=datetime.now())
        
    except Exception as e:
        flash(f'❌ Error loading medicine details: {str(e)}', 'danger')
        return redirect(url_for('view_medicines'))

@app.route('/api/reports/summary')
@login_required
def get_report_summary():
    try:
        db = get_db()
        
        # Today's summary
        today = datetime.now().date()
        
        # Today's sales
        today_sales = db.execute('''
            SELECT 
                COUNT(*) as total_sales,
                SUM(quantity_sold) as total_units,
                SUM(quantity_sold * selling_price) as total_revenue,
                AVG(quantity_sold * selling_price) as avg_transaction
            FROM sales 
            WHERE DATE(sold_on) = DATE(?)
        ''', (today,)).fetchone()
        
        # Week's sales (last 7 days)
        week_sales = db.execute('''
            SELECT 
                COUNT(*) as total_sales,
                SUM(quantity_sold) as total_units,
                SUM(quantity_sold * selling_price) as total_revenue,
                AVG(quantity_sold * selling_price) as avg_transaction
            FROM sales 
            WHERE DATE(sold_on) >= DATE(?, '-7 days')
        ''', (today,)).fetchone()
        
        # Month's sales (last 30 days)
        month_sales = db.execute('''
            SELECT 
                COUNT(*) as total_sales,
                SUM(quantity_sold) as total_units,
                SUM(quantity_sold * selling_price) as total_revenue,
                AVG(quantity_sold * selling_price) as avg_transaction
            FROM sales 
            WHERE DATE(sold_on) >= DATE(?, '-30 days')
        ''', (today,)).fetchone()
        
        # Year's sales (last 365 days)
        year_sales = db.execute('''
            SELECT 
                COUNT(*) as total_sales,
                SUM(quantity_sold) as total_units,
                SUM(quantity_sold * selling_price) as total_revenue,
                AVG(quantity_sold * selling_price) as avg_transaction
            FROM sales 
            WHERE DATE(sold_on) >= DATE(?, '-365 days')
        ''', (today,)).fetchone()
        
        # Inventory summary
        inventory_summary = db.execute('''
            SELECT 
                COUNT(DISTINCT m.id) as total_medicines,
                COUNT(b.id) as total_batches,
                SUM(b.quantity) as total_quantity,
                SUM(b.quantity * b.mrp) as stock_value,
                SUM(b.quantity * b.cost_price) as cost_value
            FROM medicines m
            LEFT JOIN batches b ON m.id = b.medicine_id
            WHERE b.expiry_date >= DATE('now') OR b.id IS NULL
        ''').fetchone()
        
        # Calculate profit margin
        potential_revenue = inventory_summary['stock_value'] or 0
        cost_value = inventory_summary['cost_value'] or 0
        profit_margin = 0
        if cost_value > 0:
            profit_margin = ((potential_revenue - cost_value) / cost_value) * 100
        
        # Expiry statistics
        expiry_stats = db.execute('''
            SELECT 
                COUNT(CASE WHEN expiry_date < DATE('now') THEN 1 END) as expired_count,
                COUNT(CASE WHEN expiry_date BETWEEN DATE('now') AND DATE('now', '+15 days') THEN 1 END) as near_expiry_count,
                COUNT(CASE WHEN expiry_date BETWEEN DATE('now', '+16 days') AND DATE('now', '+90 days') THEN 1 END) as expiring_soon_count,
                COUNT(CASE WHEN expiry_date > DATE('now', '+90 days') THEN 1 END) as good_stock_count
            FROM batches
        ''').fetchone()
        
        # Top selling medicines (last 30 days)
        top_medicines = db.execute('''
            SELECT 
                m.name,
                SUM(s.quantity_sold) as total_sold,
                SUM(s.quantity_sold * s.selling_price) as total_revenue
            FROM sales s
            JOIN batches b ON s.batch_id = b.id
            JOIN medicines m ON b.medicine_id = m.id
            WHERE DATE(s.sold_on) >= DATE(?, '-30 days')
            GROUP BY m.id, m.name
            ORDER BY total_sold DESC
            LIMIT 6
        ''', (today,)).fetchall()
        
        # Daily sales for last 7 days (for chart)
        daily_sales = db.execute('''
            SELECT 
                DATE(s.sold_on) as sale_date,
                SUM(s.quantity_sold) as total_units,
                SUM(s.quantity_sold * s.selling_price) as total_revenue
            FROM sales s
            WHERE DATE(s.sold_on) >= DATE(?, '-7 days')
            GROUP BY DATE(s.sold_on)
            ORDER BY sale_date
        ''', (today,)).fetchall()
        
        db.close()
        
        # Format the response
        return jsonify({
            'success': True,
            'summary': {
                'today': {
                    'total_sales': today_sales['total_sales'] or 0,
                    'total_units': today_sales['total_units'] or 0,
                    'total_revenue': today_sales['total_revenue'] or 0,
                    'avg_transaction': today_sales['avg_transaction'] or 0
                },
                'week': {
                    'total_sales': week_sales['total_sales'] or 0,
                    'total_units': week_sales['total_units'] or 0,
                    'total_revenue': week_sales['total_revenue'] or 0,
                    'avg_transaction': week_sales['avg_transaction'] or 0
                },
                'month': {
                    'total_sales': month_sales['total_sales'] or 0,
                    'total_units': month_sales['total_units'] or 0,
                    'total_revenue': month_sales['total_revenue'] or 0,
                    'avg_transaction': month_sales['avg_transaction'] or 0
                },
                'year': {
                    'total_sales': year_sales['total_sales'] or 0,
                    'total_units': year_sales['total_units'] or 0,
                    'total_revenue': year_sales['total_revenue'] or 0,
                    'avg_transaction': year_sales['avg_transaction'] or 0
                }
            },
            'inventory': {
                'total_medicines': inventory_summary['total_medicines'] or 0,
                'total_batches': inventory_summary['total_batches'] or 0,
                'total_quantity': inventory_summary['total_quantity'] or 0,
                'stock_value': inventory_summary['stock_value'] or 0,
                'cost_value': cost_value,
                'profit_margin': profit_margin
            },
            'expiry': {
                'expired_count': expiry_stats['expired_count'] or 0,
                'near_expiry_count': expiry_stats['near_expiry_count'] or 0,
                'expiring_soon_count': expiry_stats['expiring_soon_count'] or 0,
                'good_stock_count': expiry_stats['good_stock_count'] or 0
            },
            'top_medicines': [
                {
                    'name': med['name'],
                    'total_sold': med['total_sold'] or 0,
                    'total_revenue': med['total_revenue'] or 0
                } for med in top_medicines
            ],
            'daily_sales': [
                {
                    'date': row['sale_date'],
                    'total_units': row['total_units'] or 0,
                    'total_revenue': row['total_revenue'] or 0
                } for row in daily_sales
            ],
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        print(f"Error in report summary: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/custom')
@login_required
def get_custom_report():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({'success': False, 'error': 'Start date and end date required'}), 400
        
        db = get_db()
        
        # Sales in custom period
        sales_data = db.execute('''
            SELECT 
                COUNT(*) as total_sales,
                SUM(quantity_sold) as total_units,
                SUM(quantity_sold * selling_price) as total_revenue,
                AVG(quantity_sold * selling_price) as avg_transaction,
                COUNT(DISTINCT customer_name) as unique_customers
            FROM sales 
            WHERE DATE(sold_on) BETWEEN DATE(?) AND DATE(?)
        ''', (start_date, end_date)).fetchone()
        
        # Daily breakdown
        daily_breakdown = db.execute('''
            SELECT 
                DATE(s.sold_on) as sale_date,
                SUM(s.quantity_sold) as total_units,
                SUM(s.quantity_sold * s.selling_price) as total_revenue,
                COUNT(*) as transactions
            FROM sales s
            WHERE DATE(s.sold_on) BETWEEN DATE(?) AND DATE(?)
            GROUP BY DATE(s.sold_on)
            ORDER BY sale_date
        ''', (start_date, end_date)).fetchall()
        
        # Top medicines in period
        top_medicines = db.execute('''
            SELECT 
                m.name,
                SUM(s.quantity_sold) as total_sold,
                SUM(s.quantity_sold * s.selling_price) as total_revenue
            FROM sales s
            JOIN batches b ON s.batch_id = b.id
            JOIN medicines m ON b.medicine_id = m.id
            WHERE DATE(s.sold_on) BETWEEN DATE(?) AND DATE(?)
            GROUP BY m.id, m.name
            ORDER BY total_sold DESC
            LIMIT 10
        ''', (start_date, end_date)).fetchall()
        
        # Payment method breakdown
        payment_breakdown = db.execute('''
            SELECT 
                payment_method,
                COUNT(*) as transactions,
                SUM(quantity_sold * selling_price) as total_amount
            FROM sales 
            WHERE DATE(sold_on) BETWEEN DATE(?) AND DATE(?)
            GROUP BY payment_method
        ''', (start_date, end_date)).fetchall()
        
        db.close()
        
        return jsonify({
            'success': True,
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'summary': {
                'total_sales': sales_data['total_sales'] or 0,
                'total_units': sales_data['total_units'] or 0,
                'total_revenue': sales_data['total_revenue'] or 0,
                'avg_transaction': sales_data['avg_transaction'] or 0,
                'unique_customers': sales_data['unique_customers'] or 0
            },
            'daily_breakdown': [
                {
                    'date': row['sale_date'],
                    'total_units': row['total_units'] or 0,
                    'total_revenue': row['total_revenue'] or 0,
                    'transactions': row['transactions'] or 0
                } for row in daily_breakdown
            ],
            'top_medicines': [
                {
                    'name': med['name'],
                    'total_sold': med['total_sold'] or 0,
                    'total_revenue': med['total_revenue'] or 0
                } for med in top_medicines
            ],
            'payment_breakdown': [
                {
                    'method': row['payment_method'],
                    'transactions': row['transactions'] or 0,
                    'total_amount': row['total_amount'] or 0
                } for row in payment_breakdown
            ],
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        print(f"Error in custom report: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Helper function to create Excel report
def create_excel_report(data, headers, title, period):
    """Create an Excel workbook with the report data"""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    # Add title
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f'Period: {period}'
    ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    
    # Add headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row_data in enumerate(data, 6):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Helper function to create PDF report
def create_pdf_report(data, headers, title, period, report_type):
    """Create a PDF report with the data"""
    # Create temporary file for PDF
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdf_path = temp_file.name
    temp_file.close()
    
    # Create PDF canvas
    c = canvas.Canvas(pdf_path, pagesize=letter)
    width, height = letter
    
    # Add title
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, height - 50, title)
    
    c.setFont("Helvetica", 12)
    c.drawString(50, height - 75, f'Period: {period}')
    c.drawString(50, height - 95, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    
    # Add headers
    c.setFont("Helvetica-Bold", 10)
    y_position = height - 125
    
    # Calculate column widths
    col_width = (width - 100) / len(headers)
    
    # Draw headers
    for i, header in enumerate(headers):
        c.drawString(50 + i * col_width, y_position, header)
    
    # Draw line under headers
    c.line(50, y_position - 5, width - 50, y_position - 5)
    
    # Add data
    c.setFont("Helvetica", 9)
    y_position -= 20
    
    for row_idx, row_data in enumerate(data):
        if y_position < 50:  # New page if running out of space
            c.showPage()
            c.setFont("Helvetica", 9)
            y_position = height - 50
        
        for i, cell_value in enumerate(row_data):
            # Truncate long text
            text = str(cell_value)[:30] + "..." if len(str(cell_value)) > 30 else str(cell_value)
            c.drawString(50 + i * col_width, y_position, text)
        
        y_position -= 15
    
    # Add summary for comprehensive reports
    if report_type == 'summary':
        c.showPage()
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, height - 50, "Report Summary")
        
        c.setFont("Helvetica", 12)
        c.drawString(50, height - 100, f"Total Records: {len(data)}")
        c.drawString(50, height - 125, f"Report Type: {title}")
        c.drawString(50, height - 150, f"Generated by: Smart Pharma Assistant")
    
    c.save()
    
    # Read PDF back into BytesIO
    with open(pdf_path, 'rb') as f:
        pdf_bytes = f.read()
    
    # Clean up temp file
    os.unlink(pdf_path)
    
    return io.BytesIO(pdf_bytes)

# Updated export route with proper PDF and Excel support
@app.route('/api/reports/export')
@login_required
def export_report():
    try:
        report_type = request.args.get('type', 'summary')
        export_format = request.args.get('format', 'csv')
        period = request.args.get('period', 'today')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        print(f"Export request: type={report_type}, format={export_format}, period={period}")
        
        # Get date range
        today = datetime.now().date()
        
        if period == 'custom' and start_date and end_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            period_str = f"{start_date} to {end_date}"
        elif period == 'today':
            start_date_obj = today
            end_date_obj = today
            period_str = "Today"
        elif period == 'week':
            end_date_obj = today
            start_date_obj = end_date_obj - timedelta(days=7)
            period_str = f"Last 7 days ({start_date_obj} to {end_date_obj})"
        elif period == 'month':
            end_date_obj = today
            start_date_obj = end_date_obj - timedelta(days=30)
            period_str = f"Last 30 days ({start_date_obj} to {end_date_obj})"
        elif period == 'quarter':
            end_date_obj = today
            start_date_obj = end_date_obj - timedelta(days=90)
            period_str = f"Last 90 days ({start_date_obj} to {end_date_obj})"
        elif period == 'year':
            end_date_obj = today
            start_date_obj = end_date_obj - timedelta(days=365)
            period_str = f"Last 365 days ({start_date_obj} to {end_date_obj})"
        else:
            # Default to month
            end_date_obj = today
            start_date_obj = end_date_obj - timedelta(days=30)
            period_str = f"Last 30 days ({start_date_obj} to {end_date_obj})"
        
        db = get_db()
        
        # Fetch data based on report type
        if report_type == 'sales':
            title = f'Sales Report - {period_str}'
            data = db.execute('''
                SELECT 
                    DATE(s.sold_on) as sale_date,
                    s.customer_name,
                    s.customer_phone,
                    m.name as medicine_name,
                    b.batch_no,
                    s.quantity_sold,
                    s.selling_price,
                    s.quantity_sold * s.selling_price as total_amount,
                    s.payment_method
                FROM sales s
                JOIN batches b ON s.batch_id = b.id
                JOIN medicines m ON b.medicine_id = m.id
                WHERE DATE(s.sold_on) BETWEEN ? AND ?
                ORDER BY s.sold_on DESC
            ''', (start_date_obj, end_date_obj)).fetchall()
            
            headers = ['Date', 'Customer', 'Phone', 'Medicine', 'Batch', 'Quantity', 'Price', 'Total', 'Payment']
            rows = []
            for row in data:
                rows.append([
                    row['sale_date'],
                    row['customer_name'],
                    row['customer_phone'],
                    row['medicine_name'],
                    row['batch_no'],
                    row['quantity_sold'],
                    f"₹{row['selling_price']:.2f}",
                    f"₹{row['total_amount']:.2f}",
                    row['payment_method']
                ])
            
        elif report_type == 'inventory':
            title = f'Inventory Report - {datetime.now().strftime("%Y-%m-%d")}'
            data = db.execute('''
                SELECT 
                    m.name,
                    m.category,
                    b.batch_no,
                    b.quantity,
                    b.mrp,
                    b.cost_price,
                    b.expiry_date,
                    b.supplier,
                    b.quantity * b.mrp as stock_value
                FROM medicines m
                JOIN batches b ON m.id = b.medicine_id
                WHERE b.expiry_date >= DATE('now')
                ORDER BY m.name, b.expiry_date
            ''').fetchall()
            
            headers = ['Medicine', 'Category', 'Batch', 'Quantity', 'MRP', 'Cost', 'Expiry', 'Supplier', 'Stock Value']
            rows = []
            for row in data:
                rows.append([
                    row['name'],
                    row['category'],
                    row['batch_no'],
                    row['quantity'],
                    f"₹{row['mrp']:.2f}",
                    f"₹{row['cost_price']:.2f}",
                    row['expiry_date'],
                    row['supplier'],
                    f"₹{row['stock_value']:.2f}"
                ])
            
        elif report_type == 'expiry':
            title = f'Expiry Report - {datetime.now().strftime("%Y-%m-%d")}'
            data = db.execute('''
                SELECT 
                    m.name,
                    b.batch_no,
                    b.quantity,
                    b.mrp,
                    b.expiry_date,
                    JULIANDAY(b.expiry_date) - JULIANDAY(DATE('now')) as days_until_expiry
                FROM medicines m
                JOIN batches b ON m.id = b.medicine_id
                WHERE b.expiry_date >= DATE('now')
                ORDER BY b.expiry_date
            ''').fetchall()
            
            headers = ['Medicine', 'Batch', 'Quantity', 'MRP', 'Expiry Date', 'Days Left', 'Status']
            rows = []
            for row in data:
                days_left = int(row['days_until_expiry']) if row['days_until_expiry'] else 0
                if days_left <= 0:
                    status = 'EXPIRED'
                elif days_left <= 15:
                    status = 'URGENT (<15 days)'
                elif days_left <= 90:
                    status = 'WARNING (15-90 days)'
                else:
                    status = 'GOOD (>90 days)'
                
                rows.append([
                    row['name'],
                    row['batch_no'],
                    row['quantity'],
                    f"₹{row['mrp']:.2f}",
                    row['expiry_date'],
                    days_left,
                    status
                ])
            
        else:  # summary or comprehensive
            title = f'Comprehensive Report - {period_str}'
            
            # Get sales summary
            sales_data = db.execute('''
                SELECT 
                    DATE(s.sold_on) as sale_date,
                    s.customer_name,
                    m.name as medicine_name,
                    s.quantity_sold,
                    s.selling_price,
                    s.quantity_sold * s.selling_price as total_amount
                FROM sales s
                JOIN batches b ON s.batch_id = b.id
                JOIN medicines m ON b.medicine_id = m.id
                WHERE DATE(s.sold_on) BETWEEN ? AND ?
                ORDER BY s.sold_on DESC
                LIMIT 100
            ''', (start_date_obj, end_date_obj)).fetchall()
            
            headers = ['Date', 'Customer', 'Medicine', 'Quantity', 'Price', 'Total']
            rows = []
            for row in sales_data:
                rows.append([
                    row['sale_date'],
                    row['customer_name'],
                    row['medicine_name'],
                    row['quantity_sold'],
                    f"₹{row['selling_price']:.2f}",
                    f"₹{row['total_amount']:.2f}"
                ])
        
        db.close()
        
        # Generate file based on format
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{report_type}_{period}_{timestamp}"
        
        if export_format == 'excel':
            # Create Excel file
            excel_data = create_excel_report(rows, headers, title, period_str)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename += '.xlsx'
            response = make_response(excel_data.getvalue())
            
        elif export_format == 'pdf':
            # Create PDF file
            pdf_data = create_pdf_report(rows, headers, title, period_str, report_type)
            content_type = 'application/pdf'
            filename += '.pdf'
            response = make_response(pdf_data.getvalue())
            
        else:  # csv
            # Create CSV file
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([title])
            writer.writerow([f'Period: {period_str}'])
            writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
            writer.writerow([])
            writer.writerow(headers)
            
            # Write data
            for row in rows:
                writer.writerow(row)
            
            content_type = 'text/csv'
            filename += '.csv'
            response = make_response(output.getvalue())
        
        # Set response headers
        response.headers['Content-Type'] = content_type
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        print(f"Response headers: {dict(response.headers)}")
        print(f"Report generated: {filename}")
        return response
        
    except Exception as e:
        print(f"Error in export_report: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# Error handlers
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('500.html'), 500

if __name__ == '__main__':
    # Initialize database if not exists
    if not os.path.exists('pharma.db'):
        print("Database not found. Initializing...")
        init_database()
    
    # Start alert checker
    start_alert_checker()
    
    print("\n" + "="*50)
    print("🚀 Smart Pharma Assistant Starting...")
    print("="*50)
    print("📊 Access at: http://127.0.0.1:5000")
    print("👤 Login: admin / admin123")
    print("="*50 + "\n")
    
    app.run(debug=True, port=5000)